"""
Ryanair prijstracker: CRL <-> AOI  en  CRL <-> BLQ
Bouwt dagelijks een prijsgeschiedenis op en exporteert naar Excel.

Gebruik:
    python ryanair_scraper.py
"""

import sys, os, subprocess, time

# ── Controleer + installeer ontbrekende bibliotheken ──────────────────────────
print("=" * 56, flush=True)
print("  RYANAIR PRIJSTRACKER - Opstarten...", flush=True)
print("=" * 56, flush=True)
print(flush=True)
print("  Bibliotheken controleren...", flush=True)

for pakket, importnaam in [("requests","requests"), ("python-dateutil","dateutil"), ("openpyxl","openpyxl")]:
    try:
        __import__(importnaam)
        print(f"  [OK]  {pakket}", flush=True)
    except ImportError:
        print(f"  [!!]  {pakket} ontbreekt - wordt automatisch geinstalleerd...", flush=True)
        try:
            subprocess.check_call([sys.executable, "-m", "pip", "install", pakket],
                                  stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
            print(f"  [OK]  {pakket} geinstalleerd", flush=True)
        except Exception as e:
            print(f"  [FOUT] Kon {pakket} niet installeren: {e}", flush=True)
            input("\n  Druk op Enter om af te sluiten...")
            sys.exit(1)

print(flush=True)

import requests, json
from datetime import date, datetime
from dateutil.relativedelta import relativedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, Reference
from openpyxl.utils import get_column_letter

# ── Instellingen ───────────────────────────────────────────────────────────────
ROUTES = [
    ("CRL", "AOI", "Charleroi -> Ancona"),
    ("AOI", "CRL", "Ancona -> Charleroi"),
    ("CRL", "BLQ", "Charleroi -> Bologna"),
    ("BLQ", "CRL", "Bologna -> Charleroi"),
]
CURRENCY        = "EUR"
MONTHS_AHEAD    = 6
DEBUG           = True
WACHT_SECONDEN  = 60

# Opslagpad — pas NAS_PAD aan als je map anders heet
NAS_PAD = "/homes/admin/ryanair"   # <-- map op je NAS

if os.path.exists("/volume1"):      # Synology NAS
    DESKTOP = NAS_PAD
    os.makedirs(DESKTOP, exist_ok=True)
else:                               # Windows
    DESKTOP = os.path.join(os.path.expanduser("~"), "Desktop")
HIST_FILE = os.path.join(DESKTOP, "ryanair_geschiedenis.json")

# ── GitHub Pages publicatie ───────────────────────────────────────────────────
# Stap 1: maak gratis account op github.com
# Stap 2: maak nieuwe PUBLIC repository aan (bv. "ryanair-tracker")
# Stap 3: ga naar github.com → Settings → Developer Settings →
#          Personal Access Tokens → Generate new token (classic) → vink "repo" aan
# Stap 4: vul hieronder in en zet GITHUB_ENABLED op True

GITHUB_ENABLED   = False                # <-- zet op True na instellen
GITHUB_REPO_DIR  = "/homes/admin/ryanair"  # zelfde als NAS_PAD hierboven
GITHUB_TOKEN     = ""               # ghp_xxxxxxxxxxxx  (jouw token)
GITHUB_USERNAME  = ""               # jouw GitHub-gebruikersnaam
GITHUB_REPO_NAME = "ryanair-tracker"    # naam van je repository
GITHUB_HTML_NAME = "index.html"         # niet wijzigen

HEADERS = {
    "User-Agent": ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                   "AppleWebKit/537.36 Chrome/124.0.0.0 Safari/537.36"),
    "Referer": "https://www.ryanair.com/",
    "Accept":  "application/json",
}

# ── Kleuren ────────────────────────────────────────────────────────────────────
C = {
    "hdr_bg": "1A56DB", "hdr_fg": "FFFFFF",
    "sub_bg": "E8F0FE", "sub_fg": "1A56DB",
    "up":     "F8D7DA", "up_fg":  "721C24",
    "down":   "D4EDDA", "dn_fg":  "155724",
    "same":   "FFF3CD", "sm_fg":  "856404",
    "new":    "F0F0F0", "new_fg": "333333",
    "row_e":  "F8F9FA", "row_o":  "FFFFFF",
    "bdr":    "DEE2E6",
}

def fill(h):   return PatternFill("solid", start_color=h, end_color=h)
def fnt(h, bold=False, sz=10, u=False):
    return Font(name="Arial", color=h, bold=bold, size=sz,
                underline="single" if u else None)
def bdr():
    s = Side(style="thin", color=C["bdr"])
    return Border(left=s, right=s, top=s, bottom=s)
def aln(h="center"): return Alignment(horizontal=h, vertical="center")

# ── Logging ────────────────────────────────────────────────────────────────────
def log(msg=""):       print(msg, flush=True)
def log_ok(msg):       print(f"  [OK]   {msg}", flush=True)
def log_info(msg):     print(f"  -->    {msg}", flush=True)
def log_warn(msg):     print(f"  [!]    {msg}", flush=True)
def log_sectie(titel):
    print(flush=True)
    print("=" * 56, flush=True)
    print(f"  {titel}", flush=True)
    print("=" * 56, flush=True)

def log_stap(n, totaal, msg):
    print(f"  [{n:>2}/{totaal}]  {msg}", end="", flush=True)

# ── Geschiedenis ───────────────────────────────────────────────────────────────
def load_history():
    if os.path.exists(HIST_FILE):
        with open(HIST_FILE, encoding="utf-8") as f:
            return json.load(f)
    return {}

def save_history(hist):
    os.makedirs(os.path.dirname(HIST_FILE), exist_ok=True)
    with open(HIST_FILE, "w", encoding="utf-8") as f:
        json.dump(hist, f, ensure_ascii=False, indent=2)

# ── API ────────────────────────────────────────────────────────────────────────
def fetch_cheapest(origin, dest, year, month):
    url = (f"https://www.ryanair.com/api/farfnd/v4/oneWayFares"
           f"/{origin}/{dest}/cheapestPerDay"
           f"?outboundMonthOfDate={year}-{month:02d}-01&currency={CURRENCY}")
    try:
        r = requests.get(url, headers=HEADERS, timeout=15)
        r.raise_for_status()
        fares = r.json().get("outbound", {}).get("fares", [])
        out = {}
        for f in fares:
            dag = f.get("day", ""); pi = f.get("price")
            if not dag or f.get("soldOut") or f.get("unavailable") or pi is None:
                continue
            out[dag] = pi.get("value")
        return out
    except Exception as e:
        log_warn(f"{origin}->{dest} {year}-{month:02d}: {e}")
        return {}

def fetch_timetable(origin, dest, year, month, show_debug=False):
    url = (f"https://services-api.ryanair.com/timtbl/3/schedules"
           f"/{origin}/{dest}/years/{year}/months/{month}")
    try:
        r = requests.get(url, headers=HEADERS, timeout=15)
        r.raise_for_status()
        data = r.json()
        if show_debug and DEBUG:
            days = data.get("days", [])
            if days and days[0].get("flights"):
                log()
                log("  [DEBUG] Ruwe vluchtdata eerste vlucht:")
                log(f"  {json.dumps(days[0]['flights'][0], indent=4)}")
                log()
        schema = {}
        for de in data.get("days", []):
            dag_nr = de.get("day"); flights = []
            for f in de.get("flights", []):
                vt = _xt(f, "departureTime","time","STD","departure")
                at = _xt(f, "arrivalTime","time","STA","arrival", idx=1)
                nr = f.get("carrierCode","FR") + str(f.get("number",""))
                flights.append({"vluchtnr": nr, "vertrek": vt, "aankomst": at})
            schema[dag_nr] = flights
        return schema
    except Exception as e:
        log_warn(f"timetable {origin}->{dest} {year}-{month:02d}: {e}")
        return {}

def _xt(f, *keys, idx=0):
    for k in keys:
        v = f.get(k)
        if v is None: continue
        if isinstance(v, list):
            if len(v) > idx and v[idx]: return str(v[idx])[:5]
        elif isinstance(v, str) and v:
            return v.split("T")[1][:5] if "T" in v else v[:5]
    return "?"

# ── Scrapen ────────────────────────────────────────────────────────────────────
def scrape_and_update(hist, vandaag):
    start = date.today().replace(day=1)
    totaal_nieuw = 0

    for ri, (origin, dest, label) in enumerate(ROUTES):
        route_key = f"{origin}-{dest}"
        if route_key not in hist:
            hist[route_key] = {}

        log_sectie(f"ROUTE {ri+1}/{len(ROUTES)}:  {label}")

        log()
        log("  Stap 1/2  Prijzen ophalen...")
        log()
        prijzen = {}
        for i in range(MONTHS_AHEAD):
            t = start + relativedelta(months=i)
            log_stap(i+1, MONTHS_AHEAD, f"{t.strftime('%B %Y')}  ...")
            p = fetch_cheapest(origin, dest, t.year, t.month)
            prijzen.update(p)
            if p:
                log(f"  {len(p)} vluchten  |  goedkoopste: EUR {min(p.values()):.2f}  |  duurste: EUR {max(p.values()):.2f}")
            else:
                log("  geen vluchten")
            time.sleep(0.25)

        log()
        log("  Stap 2/2  Vertrektijden ophalen...")
        log()
        timetables = {}
        for i in range(MONTHS_AHEAD):
            t = start + relativedelta(months=i)
            log_stap(i+1, MONTHS_AHEAD, f"{t.strftime('%B %Y')}  ...")
            schema = fetch_timetable(origin, dest, t.year, t.month, show_debug=(i==0))
            if schema:
                for dag_nr, flights in schema.items():
                    datum = f"{t.year}-{t.month:02d}-{dag_nr:02d}"
                    timetables[datum] = flights
                tijden = sorted({v["vertrek"] for fl in schema.values() for v in fl if v["vertrek"] != "?"})
                log(f"  vertrektijden: {', '.join(tijden) or 'geen'}")
            else:
                log("  geen data")
            time.sleep(0.25)

        log()
        nieuw = 0
        for datum, prijs in prijzen.items():
            if datum not in hist[route_key]:
                hist[route_key][datum] = {}
            for vl in timetables.get(datum, [{}]):
                hist[route_key][datum][vandaag] = {
                    "prijs":    prijs,
                    "vluchtnr": vl.get("vluchtnr","?"),
                    "vertrek":  vl.get("vertrek","?"),
                    "aankomst": vl.get("aankomst","?"),
                }
                nieuw += 1
        totaal_nieuw += nieuw
        log_ok(f"{nieuw} metingen opgeslagen voor {label}")

    return totaal_nieuw

# ── Excel (overzicht + geschiedenis, geen per-datum-tabs) ─────────────────────
def build_excel(hist, output_path):
    wb = Workbook()
    wb.remove(wb.active)

    # ── Blad 1: Goedkoopste per dag (met AutoFilter) ───────────────────────────
    gd = wb.create_sheet("Goedkoopste per Dag", 0)
    gd.sheet_view.showGridLines = False
    gd.column_dimensions["A"].width = 2

    gd.merge_cells("B1:L1")
    gd["B1"] = "Ryanair  —  Goedkoopste vlucht per dag"
    gd["B1"].font = fnt(C["hdr_fg"], bold=True, sz=15)
    gd["B1"].fill = fill(C["hdr_bg"]); gd["B1"].alignment = aln()
    gd.row_dimensions[1].height = 34

    gd.merge_cells("B2:L2")
    gd["B2"] = (f"Laatste update: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
                f"  |  Gebruik de AutoFilter (pijltjes) om te filteren op maand of prijs"
                f"  |  Gesorteerd: goedkoopste bovenaan per route")
    gd["B2"].font = fnt("555555", sz=9)
    gd["B2"].fill = fill("EBF0FF"); gd["B2"].alignment = aln()
    gd.row_dimensions[2].height = 16

    # Kolomkoppen
    cols   = ["B","C","D","E","F","G","H","I","J","K"]
    hdrs   = ["Route","Datum","Maand","Prijs (EUR)","Vlucht","Vertrek","Aankomst","Trend","Min prijs?","# metingen"]
    widths = [26, 13, 12, 13, 9, 9, 10, 14, 11, 12]
    for cl,h,w in zip(cols,hdrs,widths):
        c = gd[f"{cl}3"]
        c.value=h; c.font=fnt(C["hdr_fg"],bold=True,sz=9)
        c.fill=fill(C["hdr_bg"]); c.alignment=aln(); c.border=bdr()
        gd.column_dimensions[cl].width=w
    gd.row_dimensions[3].height=20

    # Min prijs per route
    route_min = {}
    for origin,dest,label in ROUTES:
        key=f"{origin}-{dest}"; rh=hist.get(key,{})
        pl=[rh[d][max(rh[d].keys())]["prijs"] for d in rh if rh[d]]
        route_min[key]=min(pl) if pl else 0

    data_row = 4
    for origin,dest,label in ROUTES:
        key=f"{origin}-{dest}"; rh=hist.get(key,{})
        if not rh: continue

        rijen=[]
        for datum,vl_hist in rh.items():
            if not vl_hist: continue
            laatste=vl_hist[max(vl_hist.keys())]
            n_met=len(vl_hist)
            sorted_sds=sorted(vl_hist.keys())
            eerste=vl_hist[sorted_sds[0]]["prijs"]; laatste_p=laatste["prijs"]
            diff=laatste_p-eerste
            if len(sorted_sds)>=2:
                trend=(f"{'+'if diff>=0 else ''}EUR {diff:.2f}"
                       f" ({'omhoog' if diff>0.01 else 'omlaag' if diff<-0.01 else 'stabiel'})")
            else:
                trend="eerste meting"
            rijen.append((laatste_p,datum,laatste,n_met,trend,diff))
        rijen.sort()

        p_min_r=rijen[0][0] if rijen else 0
        p_max_r=rijen[-1][0] if rijen else 0

        for ri,(prijs,datum,info,n_met,trend,diff) in enumerate(rijen):
            maand=datum[:7]  # "2026-06"
            row_bg=C["row_e"] if ri%2==0 else C["row_o"]
            ratio=(prijs-p_min_r)/(p_max_r-p_min_r) if p_max_r>p_min_r else 0.5
            if ratio<=0.33:   cbg,cfg=C["down"],C["dn_fg"]
            elif ratio<=0.66: cbg,cfg=C["same"],C["sm_fg"]
            else:             cbg,cfg=C["up"],C["up_fg"]
            is_min=(ri==0)

            vals=[label,datum,maand,prijs,
                  info.get("vluchtnr","?"),info.get("vertrek","?"),info.get("aankomst","?"),
                  trend,"* LAAGSTE" if is_min else "",n_met]
            for ci,(cl,v) in enumerate(zip(cols,vals)):
                c=gd[f"{cl}{data_row}"]
                c.value=v; c.border=bdr(); c.alignment=aln()
                if cl=="D":
                    c.number_format="EUR #,##0.00"
                    c.fill=fill(cbg); c.font=fnt(cfg,bold=is_min,sz=10)
                elif cl=="H":
                    tbg=C["up"] if diff>0.01 else C["down"] if diff<-0.01 else C["same"]
                    tfg=C["up_fg"] if diff>0.01 else C["dn_fg"] if diff<-0.01 else C["sm_fg"]
                    c.fill=fill(tbg); c.font=fnt(tfg,bold=False,sz=9)
                elif cl=="I":
                    c.fill=fill(C["down"] if is_min else row_bg)
                    c.font=fnt(C["dn_fg"] if is_min else "999999",bold=is_min,sz=9)
                else:
                    c.fill=fill(row_bg); c.font=fnt("333333",sz=10)
            gd.row_dimensions[data_row].height=17
            data_row+=1

    # AutoFilter
    gd.auto_filter.ref = f"B3:K{data_row-1}"
    gd.freeze_panes = "B4"

    # ── Blad 2: Overzicht (geschiedenis per route) ─────────────────────────────
    ov = wb.create_sheet("Overzicht")
    ov.sheet_view.showGridLines = False
    ov.column_dimensions["A"].width = 2

    ov.merge_cells("B1:G1")
    ov["B1"] = "Ryanair Prijsgeschiedenis per route"
    ov["B1"].font = fnt(C["hdr_fg"], bold=True, sz=14)
    ov["B1"].fill = fill(C["hdr_bg"]); ov["B1"].alignment = aln()
    ov.row_dimensions[1].height = 34

    ov.merge_cells("B2:G2")
    ov["B2"] = (f"Laatste update: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
                f"  |  Open rapport.html voor interactieve grafieken en filters")
    ov["B2"].font = fnt("666666", sz=10); ov["B2"].fill = fill("EBF0FF")
    ov["B2"].alignment = aln(); ov.row_dimensions[2].height = 18

    ov_row=4
    for origin,dest,label in ROUTES:
        route_key=f"{origin}-{dest}"; route_hist=hist.get(route_key,{})
        if not route_hist: continue
        alle_scrape_datums=sorted({sd for vl in route_hist.values() for sd in vl})

        ov.merge_cells(f"B{ov_row}:G{ov_row}")
        ov[f"B{ov_row}"]=f"  {label}"
        ov[f"B{ov_row}"].font=fnt(C["sub_fg"],bold=True,sz=12)
        ov[f"B{ov_row}"].fill=fill(C["sub_bg"])
        ov[f"B{ov_row}"].alignment=aln("left"); ov.row_dimensions[ov_row].height=26
        ov_row+=1

        col_letters=[get_column_letter(i) for i in range(2,2+5+len(alle_scrape_datums))]
        headers=["Vluchtdatum","Vlucht","Vertrek","Aankomst"]+alle_scrape_datums+["Trend"]
        widths=[14,9,9,10]+[13]*len(alle_scrape_datums)+[13]
        for ci,(h,w) in enumerate(zip(headers,widths)):
            cl=col_letters[ci]; c=ov[f"{cl}{ov_row}"]
            c.value=h; c.font=fnt(C["hdr_fg"],bold=True,sz=9)
            c.fill=fill(C["hdr_bg"]); c.alignment=aln(); c.border=bdr()
            ov.column_dimensions[cl].width=w
        ov.row_dimensions[ov_row].height=20; ov_row+=1

        for ri,vlucht_datum in enumerate(sorted(route_hist.keys())):
            vl_hist=route_hist[vlucht_datum]
            if not vl_hist: continue
            sorted_sds=sorted(vl_hist.keys())
            laatste_info=vl_hist[sorted_sds[-1]]
            prijs_nu=laatste_info["prijs"]; eerste_prijs=vl_hist[sorted_sds[0]]["prijs"]
            row_bg=C["row_e"] if ri%2==0 else C["row_o"]
            if len(sorted_sds)>=2:
                diff=prijs_nu-eerste_prijs
                if diff>0.01:    trend,tbg,tfg=f"+ EUR {diff:.2f}",C["up"],C["up_fg"]
                elif diff<-0.01: trend,tbg,tfg=f"- EUR {abs(diff):.2f}",C["down"],C["dn_fg"]
                else:            trend,tbg,tfg="= gelijk",C["same"],C["sm_fg"]
            else:
                trend,tbg,tfg="eerste meting",C["new"],C["new_fg"]
            c=ov[f"{col_letters[0]}{ov_row}"]
            c.value=vlucht_datum; c.font=fnt("1A56DB",u=True,sz=10)
            c.fill=fill(row_bg); c.border=bdr(); c.alignment=aln()
            for ci2,val in enumerate([laatste_info.get("vluchtnr","?"),
                                      laatste_info.get("vertrek","?"),
                                      laatste_info.get("aankomst","?")],start=1):
                c=ov[f"{col_letters[ci2]}{ov_row}"]
                c.value=val; c.fill=fill(row_bg); c.border=bdr()
                c.alignment=aln(); c.font=fnt("333333",sz=10)
            for ci3,sd in enumerate(alle_scrape_datums):
                cl=col_letters[4+ci3]; cel=ov[f"{cl}{ov_row}"]
                cel.border=bdr(); cel.alignment=aln()
                if sd in vl_hist:
                    p=vl_hist[sd]["prijs"]; cel.value=p; cel.number_format="EUR #,##0.00"
                    prev=[x for x in sorted_sds if x<sd]
                    if prev:
                        pp=vl_hist[prev[-1]]["prijs"]
                        if p>pp+0.01:   cbg,cfg=C["up"],C["up_fg"]
                        elif p<pp-0.01: cbg,cfg=C["down"],C["dn_fg"]
                        else:           cbg,cfg=C["same"],C["sm_fg"]
                    else: cbg,cfg=row_bg,"333333"
                    cel.fill=fill(cbg); cel.font=fnt(cfg,sz=10)
                else:
                    cel.value="-"; cel.fill=fill(row_bg); cel.font=fnt("AAAAAA",sz=10)
            ct=ov[f"{col_letters[4+len(alle_scrape_datums)]}{ov_row}"]
            ct.value=trend; ct.fill=fill(tbg); ct.border=bdr()
            ct.alignment=aln(); ct.font=fnt(tfg,bold=True,sz=10)
            ov.row_dimensions[ov_row].height=18; ov_row+=1
        ov_row+=1

    wb.save(output_path)


# ── Publiceren naar GitHub Pages ──────────────────────────────────────────────
def publish_to_github(html_path: str):
    """Pusht het HTML-rapport naar GitHub Pages via git."""
    import shutil

    if not GITHUB_ENABLED:
        return
    if not all([GITHUB_REPO_DIR, GITHUB_TOKEN, GITHUB_USERNAME, GITHUB_REPO_NAME]):
        log_warn("GitHub publicatie: vul GITHUB_TOKEN, USERNAME en REPO_DIR in in het script.")
        return

    log()
    log("  GitHub Pages publiceren...")

    repo_dir  = GITHUB_REPO_DIR
    dest_html = os.path.join(repo_dir, GITHUB_HTML_NAME)

    try:
        # Kopieer nieuwe HTML naar repo
        shutil.copy2(html_path, dest_html)
        log_info(f"HTML gekopieerd naar {dest_html}")

        # Git commit & push
        remote_url = (f"https://{GITHUB_USERNAME}:{GITHUB_TOKEN}"
                      f"@github.com/{GITHUB_USERNAME}/{GITHUB_REPO_NAME}.git")

        def git(cmd):
            import subprocess
            r = subprocess.run(
                ["git"] + cmd, cwd=repo_dir,
                capture_output=True, text=True
            )
            if r.returncode != 0 and r.stderr:
                log_warn(f"git {' '.join(cmd)}: {r.stderr.strip()}")
            return r.returncode == 0

        git(["add", GITHUB_HTML_NAME])
        vandaag = date.today().isoformat()
        git(["commit", "-m", f"Ryanair prijsupdate {vandaag}"])
        git(["push", remote_url, "main"])

        url = f"https://{GITHUB_USERNAME}.github.io/{GITHUB_REPO_NAME}/"
        log_ok(f"Gepubliceerd op: {url}")

    except Exception as e:
        log_warn(f"GitHub publicatie mislukt: {e}")


# ── HTML interactief rapport ────────────────────────────────────────────────────
def build_html(hist, output_path):
    """Genereert een interactief HTML-rapport met filters en grafieken per vluchtdatum."""

    # Bouw data-structuur voor JavaScript
    js_data = {}
    for origin, dest, label in ROUTES:
        key = f"{origin}-{dest}"
        rh  = hist.get(key, {})
        route_data = []
        for datum, vl_hist in sorted(rh.items()):
            if not vl_hist: continue
            laatste_sd   = max(vl_hist.keys())
            laatste_info = vl_hist[laatste_sd]
            history = [{"scrape": sd, "prijs": vl_hist[sd]["prijs"]}
                       for sd in sorted(vl_hist.keys())]
            route_data.append({
                "datum":    datum,
                "prijs":    laatste_info["prijs"],
                "vluchtnr": laatste_info.get("vluchtnr","?"),
                "vertrek":  laatste_info.get("vertrek","?"),
                "aankomst": laatste_info.get("aankomst","?"),
                "history":  history,
                "n_metingen": len(vl_hist),
            })
        js_data[key] = {"label": label, "vluchten": route_data}

    js_str = json.dumps(js_data, ensure_ascii=False)
    update_ts = datetime.now().strftime("%d/%m/%Y %H:%M")

    html = f"""<!DOCTYPE html>
<html lang="nl">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Ryanair Prijstracker</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.0/dist/chart.umd.min.js"></script>
<style>
  *{{box-sizing:border-box;margin:0;padding:0;}}
  body{{font-family:Arial,sans-serif;background:#f0f4ff;color:#222;}}
  header{{background:#1A56DB;color:#fff;padding:18px 28px;}}
  header h1{{font-size:22px;font-weight:700;}}
  header p{{font-size:12px;opacity:.8;margin-top:4px;}}
  .container{{max-width:1400px;margin:0 auto;padding:20px;}}
  .filters{{background:#fff;border-radius:10px;padding:16px 20px;
            margin-bottom:18px;display:flex;flex-wrap:wrap;gap:14px;align-items:flex-end;
            box-shadow:0 1px 4px rgba(0,0,0,.08);}}
  .filter-group{{display:flex;flex-direction:column;gap:4px;}}
  .filter-group label{{font-size:11px;font-weight:700;color:#555;text-transform:uppercase;}}
  select,input[type=number]{{padding:7px 10px;border:1.5px solid #d0d7f0;border-radius:6px;
    font-size:13px;background:#fff;min-width:140px;}}
  select:focus,input:focus{{outline:none;border-color:#1A56DB;}}
  .btn{{padding:8px 18px;background:#1A56DB;color:#fff;border:none;border-radius:6px;
        font-size:13px;font-weight:700;cursor:pointer;}}
  .btn:hover{{background:#1447b0;}}
  .btn-reset{{background:#6c757d;}}
  .route-block{{background:#fff;border-radius:10px;margin-bottom:22px;
               box-shadow:0 1px 4px rgba(0,0,0,.08);overflow:hidden;}}
  .route-header{{padding:14px 20px;font-size:15px;font-weight:700;color:#fff;}}
  .stats-bar{{padding:8px 20px;font-size:12px;color:#555;background:#f8f9ff;
              border-bottom:1px solid #e8eaf0;}}
  table{{width:100%;border-collapse:collapse;}}
  th{{background:#1A56DB;color:#fff;padding:9px 12px;font-size:12px;
      text-align:center;position:sticky;top:0;}}
  td{{padding:8px 12px;font-size:13px;text-align:center;border-bottom:1px solid #eef0f8;}}
  tr:hover td{{background:#f0f4ff!important;cursor:pointer;}}
  .cheap{{background:#d4edda;color:#155724;}}
  .mid{{background:#fff3cd;color:#856404;}}
  .exp{{background:#f8d7da;color:#721c24;}}
  .badge-best{{background:#155724;color:#fff;border-radius:4px;padding:1px 6px;font-size:10px;}}
  .detail-row{{display:none;background:#f8faff;}}
  .detail-row.open{{display:table-row;}}
  .detail-cell{{padding:0!important;}}
  .detail-inner{{padding:16px 20px;display:flex;gap:24px;align-items:flex-start;flex-wrap:wrap;}}
  .detail-info{{min-width:220px;}}
  .detail-info h3{{font-size:14px;font-weight:700;margin-bottom:10px;color:#1A56DB;}}
  .detail-info table{{width:auto;}}
  .detail-info td{{text-align:left;padding:4px 10px 4px 0;border:none;font-size:13px;}}
  .detail-info td:first-child{{font-weight:700;color:#555;width:110px;}}
  .chart-wrap{{flex:1;min-width:300px;max-width:600px;}}
  canvas{{max-height:220px;}}
  .no-data{{text-align:center;padding:30px;color:#aaa;font-size:14px;}}
  .count-label{{font-size:11px;color:#888;margin-left:8px;}}
  .trend-up{{color:#721c24;font-weight:700;}}
  .trend-down{{color:#155724;font-weight:700;}}
  .trend-same{{color:#856404;}}
</style>
</head>
<body>
<header>
  <h1>✈ Ryanair Prijstracker</h1>
  <p>Laatste update: {update_ts} &nbsp;|&nbsp; Klik op een rij voor de prijsevolutie</p>
</header>
<div class="container">
  <div class="filters">
    <div class="filter-group">
      <label>Route</label>
      <select id="fRoute" onchange="applyFilters()">
        <option value="">Alle routes</option>
        {"".join(f'<option value="{o}-{d}">{l}</option>' for o,d,l in ROUTES)}
      </select>
    </div>
    <div class="filter-group">
      <label>Maand</label>
      <select id="fMaand" onchange="applyFilters()">
        <option value="">Alle maanden</option>
      </select>
    </div>
    <div class="filter-group">
      <label>Max prijs (EUR)</label>
      <input type="number" id="fMaxPrijs" min="0" step="5" placeholder="bv. 50"
             oninput="applyFilters()">
    </div>
    <div class="filter-group">
      <label>Min prijs (EUR)</label>
      <input type="number" id="fMinPrijs" min="0" step="5" placeholder="bv. 10"
             oninput="applyFilters()">
    </div>
    <div class="filter-group" style="flex-direction:row;gap:8px;">
      <button class="btn btn-reset" onclick="resetFilters()">Reset</button>
    </div>
    <div class="filter-group" style="margin-left:auto;align-items:flex-end;">
      <span id="countLabel" class="count-label"></span>
    </div>
  </div>
  <div id="results"></div>
</div>

<script>
const DATA = {js_str};
const KLEUREN = {{"CRL-AOI":"#1A56DB","AOI-CRL":"#E53E3E","CRL-BLQ":"#38A169","BLQ-CRL":"#D69E2E"}};
let charts = {{}};
let openRow = null;

// Vul maand-dropdown
function initMaanden() {{
  const maanden = new Set();
  for (const key in DATA) {{
    DATA[key].vluchten.forEach(v => maanden.add(v.datum.slice(0,7)));
  }}
  const sel = document.getElementById("fMaand");
  [...maanden].sort().forEach(m => {{
    const [y,mo] = m.split("-");
    const naam = new Date(+y,+mo-1).toLocaleString("nl-BE",{{month:"long",year:"numeric"}});
    sel.innerHTML += `<option value="${{m}}">${{naam}}</option>`;
  }});
}}

function applyFilters() {{
  const fRoute    = document.getElementById("fRoute").value;
  const fMaand    = document.getElementById("fMaand").value;
  const fMaxPrijs = parseFloat(document.getElementById("fMaxPrijs").value)||Infinity;
  const fMinPrijs = parseFloat(document.getElementById("fMinPrijs").value)||0;
  let totalCount  = 0;

  // Vernietig alle open grafieken
  Object.values(charts).forEach(c => c.destroy());
  charts = {{}};
  openRow = null;

  const container = document.getElementById("results");
  container.innerHTML = "";

  for (const key in DATA) {{
    if (fRoute && key !== fRoute) continue;
    const route = DATA[key];

    let vluchten = route.vluchten.filter(v => {{
      if (fMaand && !v.datum.startsWith(fMaand)) return false;
      if (v.prijs > fMaxPrijs) return false;
      if (v.prijs < fMinPrijs) return false;
      return true;
    }});
    if (!vluchten.length) continue;

    vluchten = [...vluchten].sort((a,b) => a.prijs - b.prijs);
    const pMin = vluchten[0].prijs;
    const pMax = vluchten[vluchten.length-1].prijs;
    const pAvg = vluchten.reduce((s,v)=>s+v.prijs,0)/vluchten.length;
    totalCount += vluchten.length;
    const kleur = KLEUREN[key] || "#1A56DB";

    const block = document.createElement("div");
    block.className = "route-block";
    block.innerHTML = `
      <div class="route-header" style="background:${{kleur}}">
        ${{route.label}} &nbsp;<span style="font-weight:400;font-size:13px;">(${{vluchten.length}} vluchten)</span>
      </div>
      <div class="stats-bar">
        Laagste: <b>EUR ${{pMin.toFixed(2)}}</b> &nbsp;|&nbsp;
        Gemiddeld: <b>EUR ${{pAvg.toFixed(2)}}</b> &nbsp;|&nbsp;
        Hoogste: <b>EUR ${{pMax.toFixed(2)}}</b>
      </div>
      <div style="overflow-x:auto;">
      <table>
        <thead><tr>
          <th>#</th><th>Datum</th><th>Prijs</th><th>Vlucht</th>
          <th>Vertrek</th><th>Aankomst</th><th>Trend</th><th>Metingen</th>
        </tr></thead>
        <tbody id="tbody-${{key}}"></tbody>
      </table>
      </div>`;
    container.appendChild(block);

    const tbody = document.getElementById(`tbody-${{key}}`);
    vluchten.forEach((v,idx) => {{
      const ratio = pMax>pMin ? (v.prijs-pMin)/(pMax-pMin) : 0;
      const cls   = ratio<=0.33?"cheap":ratio<=0.66?"mid":"exp";
      const isBest= idx===0;
      const hist  = v.history;
      let trend   = "—";
      let trendCls= "";
      if (hist.length>=2) {{
        const diff = hist[hist.length-1].prijs - hist[0].prijs;
        if (diff>0.01)       {{ trend=`↑ +EUR ${{diff.toFixed(2)}}`; trendCls="trend-up"; }}
        else if (diff<-0.01) {{ trend=`↓ -EUR ${{Math.abs(diff).toFixed(2)}}`; trendCls="trend-down"; }}
        else                 {{ trend="= stabiel"; trendCls="trend-same"; }}
      }}

      const rowId    = `row-${{key}}-${{idx}}`;
      const detailId = `detail-${{key}}-${{idx}}`;
      const canvasId = `chart-${{key}}-${{idx}}`;

      const tr = document.createElement("tr");
      tr.id = rowId;
      tr.style.background = isBest ? "#d4edda" : "";
      tr.innerHTML = `
        <td style="color:#aaa;font-size:11px;">${{idx+1}}</td>
        <td><b>${{v.datum}}</b> ${{isBest?'<span class="badge-best">★ LAAGSTE</span>':''}}</td>
        <td class="${{cls}}"><b>EUR ${{v.prijs.toFixed(2)}}</b></td>
        <td>${{v.vluchtnr}}</td>
        <td>${{v.vertrek}}</td>
        <td>${{v.aankomst}}</td>
        <td class="${{trendCls}}">${{trend}}</td>
        <td style="color:#888;font-size:11px;">${{v.n_metingen}} dag(en)</td>`;
      tr.onclick = () => toggleDetail(key, idx, v, kleur, detailId, canvasId, rowId);
      tbody.appendChild(tr);

      const detailTr = document.createElement("tr");
      detailTr.id = detailId;
      detailTr.className = "detail-row";
      detailTr.innerHTML = `
        <td colspan="8" class="detail-cell">
          <div class="detail-inner">
            <div class="detail-info">
              <h3>📅 ${{v.datum}} — ${{route.label}}</h3>
              <table>
                <tr><td>Vlucht</td><td>${{v.vluchtnr}}</td></tr>
                <tr><td>Vertrek</td><td>${{v.vertrek}}</td></tr>
                <tr><td>Aankomst</td><td>${{v.aankomst}}</td></tr>
                <tr><td>Huidige prijs</td><td><b>EUR ${{v.prijs.toFixed(2)}}</b></td></tr>
                <tr><td>Metingen</td><td>${{v.history.length}} dag(en)</td></tr>
                ${{v.history.length>=2 ? `<tr><td>Eerste meting</td><td>EUR ${{v.history[0].prijs.toFixed(2)}}</td></tr>
                <tr><td>Laagste ooit</td><td>EUR ${{Math.min(...v.history.map(h=>h.prijs)).toFixed(2)}}</td></tr>
                <tr><td>Hoogste ooit</td><td>EUR ${{Math.max(...v.history.map(h=>h.prijs)).toFixed(2)}}</td></tr>` : ""}}
              </table>
            </div>
            <div class="chart-wrap">
              <canvas id="${{canvasId}}"></canvas>
              ${{v.history.length<2 ? '<p style="color:#aaa;font-size:12px;margin-top:8px;">Slechts 1 meting — meer data beschikbaar na meerdere scrape-dagen.</p>' : ''}}
            </div>
          </div>
        </td>`;
      tbody.appendChild(detailTr);
    }});
  }}

  document.getElementById("countLabel").textContent =
    totalCount ? `${{totalCount}} vlucht(en) gevonden` : "";
  if (!totalCount) {{
    container.innerHTML = '<div class="no-data">Geen vluchten gevonden met deze filters.</div>';
  }}
}}

function toggleDetail(key, idx, v, kleur, detailId, canvasId, rowId) {{
  const detailTr = document.getElementById(detailId);
  const isOpen   = detailTr.classList.contains("open");

  // Sluit vorige
  if (openRow && openRow !== detailId) {{
    document.getElementById(openRow).classList.remove("open");
    const prevCanvas = document.getElementById(openRow.replace("detail-","chart-"));
    if (prevCanvas && charts[openRow]) {{ charts[openRow].destroy(); delete charts[openRow]; }}
  }}

  detailTr.classList.toggle("open", !isOpen);
  openRow = isOpen ? null : detailId;

  if (!isOpen && v.history.length >= 1) {{
    setTimeout(() => {{
      const ctx = document.getElementById(canvasId);
      if (!ctx) return;
      if (charts[detailId]) {{ charts[detailId].destroy(); }}
      charts[detailId] = new Chart(ctx, {{
        type: "line",
        data: {{
          labels: v.history.map(h => h.scrape),
          datasets: [{{
            label: `EUR prijs — ${{v.datum}}`,
            data: v.history.map(h => h.prijs),
            borderColor: kleur,
            backgroundColor: kleur + "22",
            borderWidth: 2.5,
            pointRadius: 5,
            pointHoverRadius: 7,
            fill: true,
            tension: 0.3,
          }}]
        }},
        options: {{
          responsive: true,
          plugins: {{
            legend: {{ display: false }},
            tooltip: {{
              callbacks: {{
                label: ctx => `EUR ${{ctx.parsed.y.toFixed(2)}}`
              }}
            }}
          }},
          scales: {{
            y: {{
              beginAtZero: false,
              ticks: {{ callback: v => "EUR " + v.toFixed(0) }}
            }}
          }}
        }}
      }});
    }}, 50);
  }}
}}

function resetFilters() {{
  document.getElementById("fRoute").value = "";
  document.getElementById("fMaand").value = "";
  document.getElementById("fMaxPrijs").value = "";
  document.getElementById("fMinPrijs").value = "";
  applyFilters();
}}

initMaanden();
applyFilters();
</script>
</body>
</html>"""

    with open(output_path, "w", encoding="utf-8") as f:
        f.write(html)

# ── Countdown ──────────────────────────────────────────────────────────────────
def countdown(seconden):
    log()
    log("=" * 56)
    log(f"  Klaar! Venster sluit automatisch over {seconden} seconden.")
    log("  (druk Ctrl+C om direct te sluiten)")
    log("=" * 56)
    try:
        for i in range(seconden, 0, -1):
            print(f"\r  Sluiten over {i:>3} seconden...   ", end="", flush=True)
            time.sleep(1)
        print("\r  Venster sluit nu.                 ", flush=True)
    except KeyboardInterrupt:
        log()
        log("  Handmatig gesloten.")

# ── Hoofdprogramma ─────────────────────────────────────────────────────────────
def main():
    vandaag   = date.today().isoformat()
    start_tijd = datetime.now()

    log()
    log("=" * 56)
    log("  RYANAIR PRIJSTRACKER")
    log(f"  Scrape-datum  : {vandaag}")
    log(f"  Gestart om    : {start_tijd.strftime('%H:%M:%S')}")
    log(f"  Routes        : {len(ROUTES)}")
    log(f"  Periode       : komende {MONTHS_AHEAD} maanden")
    log(f"  Geschiedenis  : {HIST_FILE}")
    log("=" * 56)

    hist = load_history()
    totaal_metingen = sum(1 for rk in hist for vd in hist[rk] for sd in hist[rk][vd])
    log()
    log_info(f"Bestaande metingen geladen: {totaal_metingen}")

    totaal_nieuw = scrape_and_update(hist, vandaag)

    log_sectie("OPSLAAN")
    save_history(hist)
    log_ok(f"Geschiedenis: {HIST_FILE}")
    log_ok(f"Nieuwe metingen vandaag: {totaal_nieuw}")

    log()
    log("  Excel rapport genereren...")
    excel_naam  = f"Ryanair_Tracker_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    excel_path  = os.path.join(DESKTOP, excel_naam)
    try:
        build_excel(hist, excel_path)
    except PermissionError:
        excel_path = os.path.join(os.path.expanduser("~"), excel_naam)
        build_excel(hist, excel_path)

    log_ok(f"Excel: {excel_path}")

    log()
    log("  HTML rapport genereren...")
    # Altijd index.html zodat GitHub Pages het direct serveert
    html_naam  = "index.html"
    html_path  = os.path.join(DESKTOP, html_naam)
    try:
        build_html(hist, html_path)
    except Exception as e_html:
        html_path = os.path.join(os.path.expanduser("~"), html_naam)
        build_html(hist, html_path)
    log_ok(f"HTML rapport: {html_path}")

    # GitHub Pages publiceren (alleen als GITHUB_ENABLED = True)
    publish_to_github(html_path)

    duur = (datetime.now() - start_tijd).seconds
    log_sectie("SAMENVATTING")
    for origin, dest, label in ROUTES:
        n = len(hist.get(f"{origin}-{dest}", {}))
        log(f"  {label:<32} {n} vluchtdatums gevolgd")
    log()
    log(f"  Looptijd   : {duur} seconden")
    log()
    log(f"  Excel staat op je BUREAUBLAD:")
    log(f"  {excel_path}")
    log()
    log(f"  HTML rapport (open in browser voor filters + grafieken):")
    log(f"  {html_path}")
    log()

    countdown(WACHT_SECONDEN)

if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        log(); log("  Gestopt.")
        sys.exit(0)
    except Exception as e:
        import traceback
        log()
        log("=" * 56)
        log("  FOUT OPGETREDEN:")
        log(f"  {e}")
        log("=" * 56)
        traceback.print_exc()
        input("\n  Druk op Enter om af te sluiten...")
